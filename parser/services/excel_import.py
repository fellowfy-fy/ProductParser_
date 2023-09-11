from dataclasses import dataclass
from io import BytesIO
from typing import Tuple

from cachetools.func import ttl_cache
from openpyxl import load_workbook

from accounts.models import CustomUser
from products.models import Category, Product


@dataclass
class ProductExtracted:
    name: str
    path: list[str]


def extract_products(contents: bytes):
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    assert ws is not None, "No active worksheet"

    path: list[Tuple[int, str]] = []
    products: list[ProductExtracted] = []
    max_level: int = 0

    for row in ws.iter_rows(min_row=2, max_col=1):

        product_name = row[0].value
        indent = row[0].alignment.indent
        if len(path) < 1 or (indent > path[-1][0] and indent > max_level):
            path.append((indent, product_name))
        elif indent < path[-1][0]:
            path.pop()
            max_level = indent
        else:
            if max_level < path[-1][0]:
                print("MAX LEVEL CHANGE")
                max_level = path[-1][0]
                del_pos, del_name = path.pop()
                products.append(ProductExtracted(name=del_name, path=[p[1] for p in path]))

            products.append(ProductExtracted(name=product_name, path=[p[1] for p in path]))

        # print(path, product_name)

    return products


@ttl_cache(ttl=60)
def get_category(name: str, author: CustomUser | None = None):
    cat, _ = Category.objects.get_or_create(name=name, defaults={"author": author})
    return cat


def create_products(products: list[ProductExtracted], author: CustomUser | None = None):
    for p in products:
        cat_name = p.path[-1]
        cat = get_category(cat_name, author=author)
        print("Product create: ", p.name)

        product, _ = Product.objects.get_or_create(
            name=p.name,
            defaults={
                "price": 0,
                "author": author,
            },
        )
        product.categories.set([cat])


def extract_import_products(contents: bytes, category: Category | None = None, author: CustomUser | None = None) -> int:
    """Extract & import products from excel. Returns imported count"""
    products = extract_products(contents)
    create_products(products, author=author)

    return len(products)
