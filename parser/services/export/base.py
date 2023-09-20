import logging
import os
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from enum import Enum
from parser.models import ParseTask
from time import time
from typing import Sequence, TypeAlias

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from accounts.models import CustomUser
from products.models import Product


@dataclass
class ReportParams:
    user: CustomUser

    filter_task: ParseTask | None = None
    date_from: date | None = None
    date_to: date | None = None
    filter_product: Product | None = None  # Filter data by product


class CellColor(Enum):
    GREEN = "FF00FF00"
    RED = "FFFF0000"


@dataclass
class CellInfo:
    value: str | None = None
    color: CellColor | str | None = None


CellValue: TypeAlias = CellInfo | str | int | float | None
CellList: TypeAlias = list[CellValue]


class BaseExcelExport:
    name = "base"  # Report name
    params: ReportParams
    book: Workbook
    sheet: Worksheet
    log: logging.Logger

    row_index: int = 0

    def __init__(self, params: ReportParams, log_level=logging.INFO) -> None:
        self.params = params
        self.log = logging.getLogger("Excel-" + self.name)
        self.log.setLevel(log_level)

    def _get_dates(self):
        """Get params dates or default"""
        start_date = self.params.date_from
        end_date = self.params.date_to
        if not end_date:
            end_date = datetime.today()
        if not start_date:
            start_date = datetime.today() - timedelta(days=15)

        return start_date, end_date

    def get_dates_list(self) -> list[date]:
        """Get list of all dates in oarams range"""
        start_date, end_date = self._get_dates()

        delta = end_date - start_date
        r = []

        for i in range(delta.days + 1):
            day = start_date + timedelta(days=i)
            r.append(day.date())

        return r

    def get_workbook(self) -> Workbook:
        workbook = Workbook()
        return workbook

    def get_report_path(self, mkdir: bool = True) -> str:
        """Get report result path"""
        path = os.path.join(settings.MEDIA_ROOT, f"reports/{self.name}.xlsx")
        if mkdir:
            os.makedirs(os.path.dirname(path), exist_ok=True)
        return path

    def sheet_init(self) -> None:
        """Init excel worksheet"""
        self.book = self.get_workbook()
        # self.sheet = self.book.create_sheet("Отчет")
        self.sheet = self.book.active

    def delete_initial_sheet(self):
        """Delete initial sheet from workbook"""
        self.book.remove_sheet(self.sheet)
        self.sheet = None

    def create_sheet(self, title: str):
        self.sheet = self.book.create_sheet(title)
        self.log.debug(f"Created sheet: {title}")

    def sheet_process(self) -> None:
        """Add data to excel worksheet"""
        raise NotImplementedError()

    def sheet_save(self) -> str:
        """Save excel worksheet"""
        report_path = self.get_report_path()
        self.book.save(report_path)
        return report_path

    def export(self) -> str:
        """Run export"""
        self.log.info(f"Starting excel export {self.name}...")

        started = time()
        self.sheet_init()
        self.sheet_process()

        self.log.info(f"Saving excel export {self.name}...")
        report_path = self.sheet_save()

        finished = time()
        time_left = round(finished - started, 2)
        self.log.info(f"Finished excel export {self.name} by {time_left} secs. Report path: {report_path}.")
        return report_path

    def extract_cell_value(self, cell_or_val: CellValue, attr="value"):
        """Extract value from cell or return raw value"""
        if isinstance(cell_or_val, CellInfo):
            return getattr(cell_or_val, attr)
        return cell_or_val

    def extract_cell_color(self, cell_or_val: CellValue):
        if isinstance(cell_or_val, CellInfo):
            return cell_or_val.color
        return None

    def insert_row(self, values: Sequence[CellValue], row_idx: int | None = None) -> None:
        self.log.debug(f"Inserting row ({row_idx}): {values}")
        values_raw = [self.extract_cell_value(v) for v in values]

        if row_idx:
            self.sheet.insert_rows(row_idx)
            for col_idx, value in enumerate(values_raw, start=1):
                cell = self.sheet.cell(row_idx, col_idx)
                cell.value = value
        else:
            self.sheet.append(values_raw)

        colors = [self.extract_cell_color(v) for v in values]
        if any(colors):
            self.set_row_colors(colors, row_idx)

    def insert_headers(self, values: list[str]) -> None:
        self.insert_row(values)

    def set_row_colors(self, colors: list[CellColor | str | None], row_idx: int | None = None) -> None:
        """If row not specified, sheet.max_row will be used."""
        if row_idx is None:
            row_idx = self.sheet.max_row

        for col_idx, color in enumerate(colors, start=1):
            if not color:
                continue
            cell = self.sheet.cell(row_idx, col_idx)

            # cell.style.font.color.index = color
            color_value = color.value if isinstance(color, CellColor) else color
            cell.font = Font(color=color_value)

    # def _get_template_path(self):
    #     return os.path.join(os.path.dirname(__file__), "excel_templates", f"report_{self.name}.xlsx")
